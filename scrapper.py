from selenium import webdriver
from bs4 import BeautifulSoup
import pandas as pd
import time
import openpyxl
import os
import yaml
import sys
from selenium.common.exceptions import NoSuchElementException

# ============================================
# All constants are defined here
# ============================================
YAML_FILE_PATH = './scrapper.yaml'
SECRET_KEY_PATH = './secret.key'

# ============================================

class DecryptPassword:  
    def __init__(self, encrypted_password):
        if not os.path.isfile('secret.key'):
            raise Exception("No secret key file found!\nPlease generate a key by "
                            "encrypting your password with the encryption program")
            sys.exit(1)

    def load_key(self):
        """
        Load the previously generated key
        """
        return open(SECRET_KEY_PATH, "rb").read()

    def decrypt_message(self, encrypted_message):
        """
        Decrypts an encrypted message
        """
        key = self.load_key()
        f = Fernet(key)
        decrypted_message = f.decrypt(encrypted_message)
        return (decrypted_message.decode())

class Scrapper:

    def __init__(self):
        try:
            with open(YAML_FILE_PATH) as yaml_config:
                self.parsed_yaml_config = yaml.load(yaml_config, Loader=yaml.FullLoader)
        except yaml.YAMLError as errmsg:
            print("Unable to read {} with error : {}".format(YAML_FILE_PATH, errmsg))
            sys.exit(1)
        try:
            self.validate_yaml()
        except Exception as errmsg:
            print("Validation of {} with error :{}".format(YAML_FILE_PATH, errmsg))
            sys.exit(1)
        self.initialize_webdriver()
        return

    def get_driver(self):
        if self.driver:
            return self.driver
        else:
            raise Exception("Driver has not been created."
                            "Please initialise the Scrapper class")

    def validate_yaml(self):
        try:
            required_list = ['url_list', 'browser_details']
            if not all(item in self.parsed_yaml_config.keys()
                       for item in required_list):
                raise Exception('YAML needs to have {}'.format(required_list))
            browser_details_rlist = ['browser_type', 'web_driver_location']
            if not all(item in self.parsed_yaml_config['browser_details'].keys()
                       for item in browser_details_rlist):
                raise Exception('The browser_details needs to have '
                                '{}'.format(browser_details_rlist))
            if not os.path.isfile(self.parsed_yaml_config['browser_details']['web_driver_location']):
                raise Exception('Cannot find webdriver file '
                                '{}'.format(self.parsed_yaml_config['browser_details']['web_driver_location']))
            url_list_rlist = ['url']
            for each_url in self.parsed_yaml_config['url_list']:
                if not all(item in each_url.keys()
                           for item in url_list_rlist):
                    raise Exception('The url_list needs to have '
                                    '{}'.format(url_list_rlist))
                if 'login_name' in each_url.keys():
                    if each_url['login_name'] is not None:
                        if ('login_name_element' not in each_url.keys() or 
                            each_url['login_name_element'] is None):
                            raise Exception("login_name_element is required if "
                                            "login_name is defined in a url_list")
                if 'login_password' in each_url.keys():
                    if each_url['login_password'] is not None:
                        if ('login_password_hashed' not in each_url.keys() or 
                            each_url['login_password_hashed'] is None):
                            raise Exception("login_password_hashed is required if "
                                            "login_password is defined in a url_list")
                        if ('login_password_element' not in each_url.keys() or 
                            each_url['login_password_element'] is None):
                            raise Exception("login_password_element is required if "
                                            "login_password is defined in a url_list")
        except Exception as errmsg:
            raise Exception(errmsg)
        return

    def initialize_webdriver(self):
        try:
            suported_browser = ['firefox', 'chrome']
            input_bowser_type = str(self.parsed_yaml_config['browser_details']['browser_type'])
            if 'firefox' in input_bowser_type.lower():
                self.driver = webdriver.Firefox(self.parsed_yaml_config['browser_details']['web_driver_location'])
                time.sleep(5)
            elif 'chrome' in input_bowser_type.lower():
                self.driver = webdriver.Chrome(self.parsed_yaml_config['browser_details']['web_driver_location'])
                time.sleep(5)
            else:
                print("{} is not supported."
                      "\nProgram only supports browser_type : "
                      "{}".format(input_bowser_type, suported_browser))
                sys.exit(1)
            if self.driver:
                self.driver.maximize_window()
        except Exception as errmsg:
            print("Error in initialize_webdriver: {}".format(errmsg))
            sys.exit(1)
        return

    def __del__(self):
        if self.driver:
            self.driver.close()

    def unhash_password(self, encrypted_pass):
        try:
            decrypt_obj = DecryptPassword()
            decrypted_pass = decrypt_obj.decrypt_message(encrypted_pass)
            return decrypted_pass
        except Exception as errmsg:
            print("Unable to unhash password with error: {}".format(errmsg))
            sys.exit(1)

    def login(self, url_obj):
        try:
            if 'login_name' in url_obj.keys():
                if url_obj['login_name'] is not None:
                    login_name_element = url_obj['login_name_element']
                    login_name = url_obj['login_name']
                    try:
                        inputElement = self.driver.find_element_by_name(login_name_element)
                        inputElement.clear()
                        inputElement.send_keys(login_name)
                    except NoSuchElementException:
                        print("Did not find the login_name_element: "
                              "{}".format(login_name_element))
                    except Exception as errmsg:
                        print("Error in login: {}".format(errmsg))
            if 'login_password' in url_obj.keys():
                if url_obj['login_password'] is not None:
                    login_password_element = url_obj['login_password_element']
                    login_password = url_obj['login_password']
                    if url_obj['login_password_hashed']:
                        login_password = self.unhash_password()
                    try:
                        inputElement = self.driver.find_element_by_name(login_password_element)
                        inputElement.clear()
                        inputElement.send_keys(login_password)
                    except NoSuchElementException:
                        print("Did not find the login_password_element: "
                              "{}".format(login_password_element))
                    except Exception as errmsg:
                        print("Error in login: {}".format(errmsg))

            if inputElement:
                inputElement.submit()
                time.sleep(5)
                return True
            else:
                print("No login required")
                return False
        except Exception as errmsg:
            raise Exception("Login Failed with error: {}".format(errmsg))

    def get_page(self, url_obj):
        try:
            url = url_obj['url']
            self.driver.get(url)
            time.sleep(5)
            return True
        except Exception as errmsg:
            raise Exception("Error getting page {}: {}".format(url, errmsg))
        return False

    def go_to_page(self, url_num):
        try:
            url_list = self.parsed_yaml_config['url_list']
            url_obj = url_list[int(url_num) - 1]
            self.get_page(url_obj)
            self.login(url_obj)
        except Exception as errmsg:
            raise Exception("Error: {}".format(errmsg))
        return

    def convert_to_excel(self, data_list, dest_path):
        """
        data_list should be a list of dictionaries
        """
        try:
            headings = data_list[0].keys()
            from openpyxl import Workbook
            workbook = Workbook()
            sheet = workbook.active
            # Write headings
            row_num = 1
            col_num = 1
            for heading in headings:
                sheet.cell(row = row_num, column = col_num).value = heading
                col_num += 1
            col_num = 1
            row_num = 2
            # Write data
            for each in data_list:
                for heading in headings:
                    sheet.cell(row = row_num, column = col_num).value = each[heading]
                    col_num += 1
                col_num = 1
                row_num += 1
            workbook.save(dest_path)
        except Exception as errmsg:
            raise("Error in convert_to_excel: {}".format(errmsg))